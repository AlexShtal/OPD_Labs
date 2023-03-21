from bs4 import BeautifulSoup
import openpyxl
import requests

url = "https://omsk.hh.ru/search/vacancy?text=Python&area=68"
headers = {'User-Agent': 'Mozilla/5.0 (Linux; Android 5.1.1; SM-G928X Build/LMY47X) AppleWebKit/537.36 '
                         '(KHTML, like Gecko) Chrome/47.0.2526.83 Mobile Safari/537.36'}

page = requests.get(url, headers=headers)
print(page.status_code)

all_vacancies = []
vac_names = []
vac_refs = []
vac_salaries = []
vac_companies = []
data = [vac_names, vac_refs, vac_salaries, vac_companies]

soup = BeautifulSoup(page.text, "html.parser")

all_vacancies = soup.findAll("div", class_="vacancy-serp-item-body__main-info")

for vac in all_vacancies:
    # Parse names
    name = vac.find("h3", class_="bloko-header-section-3").find("a", class_="serp-item__title")
    vac_names.append(name.text)

    # Parse refs
    ref = vac.find("a", class_="serp-item__title").get("href")
    vac_refs.append(ref)

    # Parse salaries
    salary = vac.find("span", class_="bloko-header-section-3")
    if salary:
        vac_salaries.append(salary.text)
    else:
        vac_salaries.append("-")

    # Parse companies
    company = vac.find("div", class_="vacancy-serp-item-company").find("a",
                                                                       class_="bloko-link bloko-link_kind-tertiary")
    vac_companies.append(company.text)

for i in range(len(vac_companies)):
    print("{0}\t{1}\t{2}\t{3}".format(vac_names[i], vac_refs[i], vac_salaries[i], vac_companies[i]))

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Vacancies"

column_names = ["Вакансия", "Ссылка", "Заработная плата", "Компания"]

for row in range(1, 21):
    for column in range(1, 5):
        if row == 1:
            sheet.cell(row=row, column=column, value=column_names[column - 1])
        else:
            sheet.cell(row=row, column=column, value=data[column - 1][row - 1])

wb.save('Python_vacs.xlsx')
